import os.path
from typing import Callable, Any

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')  # Set the backend to 'Agg'
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO
from collections import defaultdict
from scipy.stats import binned_statistic

import metric_definitions
from gt_pred_set import GtPredSet

class EvalReport:

    not_available_marker = "--"


    def __init__(self, gt_pred_set: GtPredSet, report_save_dir: str):
        self.gt_pred_set = gt_pred_set
        self.report_save_dir = report_save_dir
        self.report = {}
        self.figures = {}
        self.wb = Workbook()
        self.ws = self.wb.active
        self._setup_excel_styles()

    def _setup_excel_styles(self):
        self.heading_font = Font(bold=True, size=12)
        self.heading_alignment = Alignment(horizontal='center')
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))


    def _add_error(self, name: str, error_function: Callable[[list, list], Any]):
        errors = {}
        class_results = []
        for object_class in self.gt_pred_set.get_class_names():
            gt, pred = self.gt_pred_set.get_gt_pred_lists(object_class)
            if len(gt) == 0:
                error_value = self.not_available_marker
            else:
                error_value = error_function(gt, pred)
            errors[object_class] = error_value
            class_results.append(error_value)

        gt, pred = self.gt_pred_set.get_gt_pred_lists()
        errors["All (mean object-wise)"] = error_function(gt, pred)

        errors["All (mean class-wise)"] = np.mean(class_results)

        self.report[name] = errors
        self._add_metric_to_excel(name, errors)

    def add_abs_relative_error(self):
        self._add_error('Abs. Relative Error', metric_definitions.aggregate_mean_abs_rel)
        return self

    def add_rmse(self):
        self._add_error('RMSE', metric_definitions.aggregate_rmse)
        return self

    def add_a1_threshold_accuracy(self):
        self._add_error('A1: Percent δ < 1,25', metric_definitions.aggregate_a1_threshold)
        return self




    def _add_error_stddev_var(self, name: str, error_function: Callable[[list, list, bool], Any]):
        errors = {'metric': {}, 'stddev': {}, 'variance': {}}
        class_results_metric = []
        class_results_stddev = []
        class_results_var = []

        for object_class in self.gt_pred_set.get_class_names():
            gt, pred = self.gt_pred_set.get_gt_pred_lists(object_class)
            if len(gt) == 0:
                errors['metric'][object_class] = self.not_available_marker
                errors['stddev'][object_class] = self.not_available_marker
                errors['variance'][object_class] = self.not_available_marker
            else:
                metric, stddev, var = error_function(gt, pred, stddev_var=True)
                errors['metric'][object_class] = metric
                errors['stddev'][object_class] = stddev
                errors['variance'][object_class] = var
                class_results_metric.append(metric)
                class_results_stddev.append(stddev)
                class_results_var.append(var)

        gt, pred = self.gt_pred_set.get_gt_pred_lists()
        metric, stddev, var = error_function(gt, pred, stddev_var=True)
        errors['metric']["All (mean object-wise)"] = metric
        errors['stddev']["All (mean object-wise)"] = stddev
        errors['variance']["All (mean object-wise)"] = var

        errors['metric']["All (mean class-wise)"] = np.mean(class_results_metric)
        errors['stddev']["All (mean class-wise)"] = np.std(class_results_stddev)
        errors['variance']["All (mean class-wise)"] = np.var(class_results_var)

        self.report[name] = errors
        self._add_metric_to_excel_stddev_var(name, errors)

    def add_abs_relative_error_stddev_var(self):
        self._add_error_stddev_var('Abs. Relative Error', metric_definitions.aggregate_mean_abs_rel)
        return self

    def add_rmse_stddev_var(self):
        self._add_error_stddev_var('RMSE', metric_definitions.aggregate_rmse)
        return self

    def add_a1_threshold_accuracy_stddev_var(self):
        self._add_error_stddev_var('A1: Percent δ < 1,25', metric_definitions.aggregate_a1_threshold)
        return self




    def add_relative_error_binned(self, bins=[0, 5000, 10000, 15000, 20000, 25000, 30000, 35000]):
        self._check_data_integrity()
        class_bins = {}

        for object_class in self.dataset:
            binned_errors = {f'{bin_start}-{bin_end}': [] for bin_start, bin_end in zip(bins[:-1], bins[1:])}
            for gt, pred in self.dataset[object_class]:
                for bin_start, bin_end in zip(bins[:-1], bins[1:]):
                    if bin_start <= gt < bin_end:
                        error = self.relative_error(gt, pred)
                        binned_errors[f'{bin_start}-{bin_end}'].append(error)
                        break

            # Calculate mean error for each bin
            class_bins[object_class] = {bin_range: np.mean(errors) if errors else None for bin_range, errors in
                                        binned_errors.items()}

        bin_ranges = [f'{bin_start}-{bin_end}' for bin_start, bin_end in zip(bins[:-1], bins[1:])]
        self._add_binned_metric_to_excel('Binned Relative Error', class_bins, bin_ranges)
        self.report['Binned Relative Error'] = class_bins
        return self

    def _add_binned_error(self, name: str, error_function: Callable[[list, list], Any], bins=[0, 5000, 10000, 15000, 20000, 25000, 30000, 35000]):
        class_bins = {}

        for object_class in self.gt_pred_set.get_class_names():
            for bin_start, bin_end in zip(bins[:-1], bins[1:]):
                gt_binned, pred_binned = self.gt_pred_set.get_gt_pred_lists(object_class, gt_bin_min=bin_start, gt_bin_max=bin_end)
                if len(gt_binned) == 0:
                    error_value_bin = self.not_available_marker
                else:
                    error_value_bin = error_function(gt_binned, pred_binned)

                if object_class not in class_bins:
                    class_bins[object_class] = {}
                class_bins[object_class][f'{bin_start}-{bin_end}'] = error_value_bin

        bin_ranges = [f'{bin_start}-{bin_end}' for bin_start, bin_end in zip(bins[:-1], bins[1:])]
        self._add_binned_metric_to_excel(name, class_bins, bin_ranges)
        self.report[name] = class_bins



    def add_abs_relative_error_binned(self, bins=[0, 5000, 10000, 15000, 20000, 25000, 30000, 35000]):
        self._add_binned_error("Abs. Relative Error Binned", metric_definitions.aggregate_mean_abs_rel, bins)
        return self

    def add_abs_error_binned(self, bins=[0, 5000, 10000, 15000, 20000, 25000, 30000, 35000]):
        self._add_binned_error("Abs. (metric) Error Binned", metric_definitions.aggregate_mean_abs_metric, bins)
        return self

    def add_rmse_binned(self, bins=[0, 5000, 10000, 15000, 20000, 25000, 30000, 35000]):
        self._add_binned_error("RMSE Binned", metric_definitions.aggregate_rmse, bins)
        return self

    def add_a1_threshold_binned(self, bins=[0, 5000, 10000, 15000, 20000, 25000, 30000, 35000]):
        self._add_binned_error("A1: Percent δ < 1,25 Binned", metric_definitions.aggregate_a1_threshold, bins)
        return self

    # def add_log_rmse(self):
    #     self._check_data_integrity()
    #     log_rmse_values = {}
    #     for object_class, measurements in self.dataset.items():
    #         errors = [self.log_rmse_error(gt, pred) for gt, pred in measurements if gt > 0 and pred > 0]
    #         log_rmse_values[object_class] = np.sqrt(np.mean(errors))
    #     self.report['Log RMSE'] = log_rmse_values
    #     self._add_metric_to_excel('Log RMSE', log_rmse_values)
    #     return self


    def _add_metric_to_excel(self, metric_name, values):
        row_num = self.ws.max_row + 1
        self.ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(values) + 1)
        cell = self.ws.cell(row=row_num, column=1)
        cell.value = metric_name
        cell.font = self.heading_font
        cell.alignment = self.heading_alignment

        classes = list(values.keys())
        self.ws.append(classes)
        self.ws.append([values[cls] for cls in classes])

        # Add empty row after each metric bloc
        self.ws.append([""])

    def _add_metric_to_excel_stddev_var(self, metric_name, values):
        row_num = self.ws.max_row + 1

        # Add metric name header
        self.ws.merge_cells(start_row=row_num, start_column=1,
                            end_row=row_num, end_column=len(values['metric']) + 2)  # +2 for the label column
        cell = self.ws.cell(row=row_num, column=1)
        cell.value = metric_name
        cell.font = self.heading_font
        cell.alignment = self.heading_alignment

        # Add column headers (classes)
        classes = list(values['metric'].keys())
        header_row = [""] + classes  # Empty cell for label column
        self.ws.append(header_row)

        # Add metric values in three rows with labels
        metric_row = ["Metric"] + [values['metric'][cls] for cls in classes]
        stddev_row = ["StdDev"] + [values['stddev'][cls] for cls in classes]
        variance_row = ["Variance"] + [values['variance'][cls] for cls in classes]

        self.ws.append(metric_row)
        self.ws.append(stddev_row)
        self.ws.append(variance_row)

        # Add empty row after each metric block
        self.ws.append([""])

    def _add_binned_metric_to_excel(self, metric_name, values, bins):
        row_num = self.ws.max_row + 1
        self.ws.merge_cells(start_row=row_num, start_column=2, end_row=row_num, end_column=len(bins) + 1)
        cell = self.ws.cell(row=row_num, column=2)
        cell.value = metric_name
        cell.font = self.heading_font
        cell.alignment = self.heading_alignment
        row_num += 1

        # Schreibe Bins als Spaltenüberschriften, beginnend bei der zweiten Spalte
        for col_num, bin_range in enumerate(bins, start=2):
            cell = self.ws.cell(row=row_num, column=col_num)
            cell.value = bin_range
            cell.font = self.heading_font
            cell.border = self.thin_border
            cell.alignment = self.heading_alignment
        row_num += 1

        # Schreibe Klassenfehler für jeden Bin, beginnend bei der zweiten Spalte
        for object_class, bins_dict in values.items():
            self.ws.cell(row=row_num, column=1).value = object_class
            for col_num, (bin_range, error) in enumerate(bins_dict.items(), start=2):
                self.ws.cell(row=row_num, column=col_num).value = error if error is not None else 'N/A'
            row_num += 1

        # Füge eine leere Zeile nach jedem Metrikblock hinzu
        self.ws.append([""])

    def write_to_excel(self, filename):
        filepath = os.path.join(self.report_save_dir, filename)
        self.wb.save(filepath)







